# Módulo | Análise de Dados: Coleta de Dados I
# Caderno de Exercícios

# Tópicos
# 1.Arquivos CSV;
# 2.Arquivos Texto;
# 3.Arquivos Excel.

# Vamos explorar dados de crédito presentes no arquivo credito.xlsx. Os dados estão no formato de Excel (XLSX) e contém informações sobre clientes de uma instituição financeira. Em especial, estamos interessados em explicar a segunda coluna, chamada de default, que indica se um cliente é adimplente (default = 0), ou inadimplente (default = 1), ou seja, queremos entender o porque um cliente deixa de honrar com suas dívidas baseado no comportamento de outros atributos, como salário, escolaridade e movimentação financeira.

# 1. Excel para CSV
# Utilizando o pacote Python openpyxl, extraia os seguintes as colunas id, sexo e idade para dos clientes inadimplentes (default = 1) e solteiros (estado_civil = 'solteiro'). Salves os dados extraídos no arquivo csv credito.csv separado por ;.

# Extrando o id, sexo e idade dos usuários de acordo com a coluna inadimplentes e se são solteiros ou não.

import csv # uso do pacote csv para converter arquivos xlsx em csv
from openpyxl import load_workbook # importando o pacote load para ler o arquivo xlsx

workbook = load_workbook('credito.xlsx') # leitura do arquivo
planilha = workbook.active # salvando os dados do arquivo na variável planilha

ids = []  # criação das listas vazias para armazenamento dos dados coletados
idades = []
sexos = []

for linha in planilha.iter_rows(min_row=1, values_only=True): # leitura de linha por linha para coletar os dados verdadeiros, pulando o cabeçalho
  id, idade, sexo, estado_civil, default = linha[0], linha[2], linha[3], linha[6], linha[1]

  if default == 1 and estado_civil.lower() == 'solteiro': # salvando os dados nas recpectivas listas
    ids.append(id)
    idades.append(idade)
    sexos.append(sexo)

with open(file='./credito.csv', mode='w', encoding='utf8', newline='') as arquivo: # abrindo o arquivo e escrevendo os dados coletados acima separados por ";"
  escritor_csv = csv.writer(arquivo, delimiter=';')
  escritor_csv.writerow(['id', 'idade', 'sexo'])

  for i in range(len(ids)): #
    escritor_csv.writerow([ids[i], idades[i], sexos[i]])

# 2. Excel para JSON
# Como preparação para o próximo módulo, vamos trabalhar com o JSON, um formato semi-estruturado, muito utilizado em transmissão de dados da web e equivalente a um dicionário Python.
# Utilizando o pacote Python openpyxl visto em aula, extraia os dados das colunas escolaridade e tipo_cartao, removendo duplicados. Com os dados, construa o dicionário Python credito.

escoladirade = []

cabecalho = next(planilha.values)
indice_escolaridade = cabecalho.index('escolaridade')
escolaridade = [linha[indice_escolaridade] for linha in planilha.values if linha[indice_escolaridade] != 'escolaridade']

print(set(list(escolaridade)))
print(type(escolaridade))

tipo_cartao = []

cabecalho = next(planilha.values)
indice_tipo_cartao = cabecalho.index('tipo_cartao')
tipo_cartao = [linha[indice_tipo_cartao] for linha in planilha.values if linha[indice_tipo_cartao] != 'tipo_cartao']

print(set(list(tipo_cartao)))
print(type(tipo_cartao))

credito = {
    'tipo_cartao': ['blue', 'platinum', 'gold', 'silver'],
    'escolaridade': ['mestrado', 'na', 'sem educacao formal', 'doutorado', 'ensino medio', 'graduacao']
}

print(type(credito))

import json

credito_json = json.dumps(credito, indent=4)
print(credito_json)

# 3. BÔNUS: Texto para CSV
# No arquivo de texto ebac.txt você encontra o texto presente na pasta de arquivos.
# Extraia os números de contato do arquivo texto ebac.txt e salve-os no arquivo csv contato_ebac.csv com o separador ;.

import re

with open(file='./ebac.txt', mode='r', encoding='utf8') as arquivo:
  texto = arquivo.read()

contatos_extraidos = re.findall('\+\d{2}\s\(\d{2}\)\s\d{4}-\d{4}', texto)
print(contatos_extraidos)

padrao_contato = r'(?:WhatsApp|Telefone)'

tipos_contato = re.findall(padrao_contato, texto)

print(tipos_contato)

import csv

contatos = list(zip(contatos_extraidos, tipos_contato))

with open(file='./contato_ebac.csv', mode='w', encoding='utf8') as arquivo:
  escritor_csv = csv.writer(arquivo, delimiter=';')
  escritor_csv.writerow(['numero', 'tipo'])

  for contato in contatos:
    escritor_csv.writerow(contato)